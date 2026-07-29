import openpyxl
import urllib.request
import json
import os
import re

TBCA_RAW_URL = "https://raw.githubusercontent.com/raul-rznd/web-scraping-tbca/main/alimentos.txt"

def parse_val(val):
    if val is None or val == 'NA' or val == 'Tr' or val == 'tr' or val == '-':
        return 0.0
    try:
        val_str = str(val).replace(',', '.').strip()
        return round(float(val_str), 1)
    except (ValueError, TypeError):
        return 0.0

def parse_taco_xlsx(xlsx_path):
    """Extrai os 597 alimentos da Tabela TACO (UNICAMP 4ª Edição) do arquivo Excel."""
    print(f"Lendo arquivo Excel oficial TACO 4ª Edição: {xlsx_path}...")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    sheet = wb['CMVCol taco3']

    current_category = 'Geral'
    taco_items = []
    seen_names = set()

    for row_idx in range(4, sheet.max_row + 1):
        c1 = sheet.cell(row_idx, 1).value
        c2 = sheet.cell(row_idx, 2).value

        # Header detection
        if c1 is not None and isinstance(c1, str) and not c1.isdigit() and c2 is None:
            cat_candidate = c1.strip()
            if cat_candidate.lower() not in ['número do', 'número do alimento', 'legenda']:
                current_category = cat_candidate

        # Data Row Detection
        elif isinstance(c1, int) or (isinstance(c1, str) and c1.isdigit()):
            food_id = int(c1)
            name = str(c2).strip() if c2 else ''
            
            if not name or name in seen_names:
                continue

            kcal = parse_val(sheet.cell(row_idx, 4).value)
            protein = parse_val(sheet.cell(row_idx, 6).value)
            fat = parse_val(sheet.cell(row_idx, 7).value)
            carbs = parse_val(sheet.cell(row_idx, 9).value)
            fiber = parse_val(sheet.cell(row_idx, 10).value)

            taco_items.append({
                'name': name,
                'category': current_category,
                'energy_kcal': kcal,
                'protein_g': protein,
                'carbs_g': carbs,
                'fat_g': fat,
                'fiber_g': fiber,
                'source': 'TACO'
            })
            seen_names.add(name)

    print(f"TACO UNICAMP extraído com sucesso! Total de alimentos: {len(taco_items)}")
    return taco_items

def parse_tbca_usp():
    """Extrai os 5.643 alimentos da TBCA (USP 5.0)."""
    print("Baixando acervo completo da TBCA (USP)...")
    headers = {'User-Agent': 'MacroNutrition/1.0'}
    
    try:
        req = urllib.request.Request(TBCA_RAW_URL, headers=headers)
        with urllib.request.urlopen(req, timeout=15) as resp:
            content = resp.read().decode('utf-8')
            lines = content.splitlines()
            
            parsed_foods = []
            seen_names = set()

            for line in lines:
                if not line.strip():
                    continue
                try:
                    item = json.loads(line)
                    raw_name = item.get('descricao', '').strip()
                    raw_name = re.sub(r',\s*$', '', raw_name)
                    category = item.get('classe', 'Geral').strip()

                    if not raw_name or raw_name in seen_names:
                        continue

                    energy_kcal = 0.0
                    protein_g = 0.0
                    carbs_g = 0.0
                    fat_g = 0.0
                    fiber_g = 0.0

                    nutrientes = item.get('nutrientes', [])
                    for n in nutrientes:
                        comp = n.get('Componente', '').strip()
                        unit = n.get('Unidades', '').strip()
                        val = parse_val(n.get('Valor por 100g'))

                        if comp == 'Energia' and unit == 'kcal':
                            energy_kcal = val
                        elif 'Proteína' in comp:
                            protein_g = val
                        elif 'Carboidrato' in comp:
                            carbs_g = val
                        elif 'Lipídeo' in comp or 'Lipídios' in comp:
                            fat_g = val
                        elif 'Fibra' in comp:
                            fiber_g = val

                    parsed_foods.append({
                        'name': raw_name,
                        'category': category,
                        'energy_kcal': energy_kcal,
                        'protein_g': protein_g,
                        'carbs_g': carbs_g,
                        'fat_g': fat_g,
                        'fiber_g': fiber_g,
                        'source': 'TBCA'
                    })
                    seen_names.add(raw_name)

                except Exception:
                    continue

            print(f"TBCA USP extraído com sucesso! Total de alimentos: {len(parsed_foods)}")
            return parsed_foods
    except Exception as e:
        print(f"Erro ao baixar TBCA USP: {e}")
        return []

def main():
    base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    xlsx_path = os.path.join(base_dir, 'Taco-4a-Edicao.xlsx')
    target_json = os.path.join(base_dir, 'app', 'data', 'foods_database.json')

    taco_foods = parse_taco_xlsx(xlsx_path)
    tbca_foods = parse_tbca_usp()

    combined_foods = taco_foods + tbca_foods
    print(f"Total unificado: {len(combined_foods)} alimentos (TACO UNICAMP + TBCA USP)!")

    with open(target_json, 'w', encoding='utf-8') as f:
        json.dump(combined_foods, f, ensure_ascii=False, indent=2)

    print(f"Arquivo {target_json} atualizado com sucesso!")

if __name__ == '__main__':
    main()
